from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import TrainingPlan, WEEKDAYS


GREEN = "1F4E46"
LIGHT_GREEN = "F3F6F4"
TEXT = "1F2933"
WHITE = "FFFFFF"
BORDER = "D5DDD8"


def export_plan_xlsx(plan: TrainingPlan, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Training Plan"
    _set_layout(sheet)
    _write_header(sheet, plan)
    _write_weeks(sheet, plan)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def _set_layout(sheet) -> None:
    widths = {
        "A": 1.5,
        "B": 10.5,
        "C": 16,
        "D": 13,
        "E": 13,
        "F": 13,
        "G": 13,
        "H": 13,
        "I": 13,
        "J": 8,
        "K": 2,
        "L": 10,
        "M": 30,
        "N": 1.5,
        "O": 10,
        "P": 22,
    }
    for col, width in widths.items():
        sheet.column_dimensions[col].width = width
    for row in range(1, 260):
        sheet.row_dimensions[row].height = 26
    sheet.sheet_view.showGridLines = False


def _write_header(sheet, plan: TrainingPlan) -> None:
    profile = plan.profile
    rows = [
        (plan.plan_title, "", "", "", "", "", "", "", "", "", "Race", profile.race_name, "", "", "", ""),
        ("Start", profile.start_date, "", "Race Day", profile.race_date, "", "Distance", plan.race_label, "", "", "Goal", profile.goal_description, "", "", "", ""),
        (
            "Current baseline",
            f"{profile.current_weekly_km:g} km/week; {profile.longest_recent_run_km:g} km long run",
            "",
            "Goal pace",
            plan.goal_pace_per_km or "Not set",
            "",
            "Current pace",
            profile.current_marathon_pace or "Not set",
            "",
            "",
            "Primary risks",
            profile.primary_risks,
            "",
            "",
            "",
            "",
        ),
        (
            "Default week shape",
            f"{profile.workout_day} workout, {profile.long_run_day} long run, strength {', '.join(profile.strength_days)}",
            "",
            "Race specifics",
            profile.race_specifics,
            "",
            "Fuel",
            profile.fuel_notes,
            "",
            "",
            "Admin",
            profile.admin_notes,
            "",
            "",
            "",
            "",
        ),
        ("Use", "Enter actual run notes in the Actual, Distance, and Remarks rows. Planned totals calculate from day columns.", "", "", "", "", "", "", "", "", "", "", "", "", "", ""),
    ]
    for row_index, values in enumerate(rows, start=1):
        for col_index, value in enumerate(values, start=2):
            cell = sheet.cell(row_index, col_index, value)
            _body(cell)
    for row in range(1, 6):
        sheet.row_dimensions[row].height = 46 if row != 1 else 80
    for cell_ref in ["B1", "K1", "B2", "E2", "H2", "K2", "B3", "E3", "H3", "K3", "B4", "E4", "H4", "K4", "B5"]:
        _label(sheet[cell_ref])

    sheet["B7"] = "Plan KPI"
    sheet["C7"] = "Formula"
    sheet["E7"] = "Value"
    sheet["G7"] = "Notes"
    for cell_ref in ["B7", "C7", "E7", "G7"]:
        _label(sheet[cell_ref], light=True)
    sheet["B8"] = "Training weeks"
    sheet["E8"] = len(plan.weeks)
    sheet["G8"] = f"{profile.start_date:%d %b} to {profile.race_date:%d %b} inclusive"
    sheet["B9"] = "Peak planned week"
    sheet["E9"] = max(week.target_km for week in plan.weeks)
    sheet["G9"] = "Based on weekly planned distance totals"
    for row in range(8, 10):
        for col in range(2, 8):
            _body(sheet.cell(row, col))


def _write_weeks(sheet, plan: TrainingPlan) -> None:
    for index, week in enumerate(plan.weeks):
        start_row = 10 + index * 10
        date_row = start_row
        week_row = start_row + 1
        type_row = start_row + 2
        plan_row = start_row + 3
        planned_distance_row = start_row + 4
        actual_row = start_row + 5
        actual_distance_row = start_row + 6
        remarks_row = start_row + 7

        sheet.cell(date_row, 2, f"Week {week.week_number}")
        _label(sheet.cell(date_row, 2))
        for offset, day in enumerate(WEEKDAYS):
            col = 3 + offset
            sheet.cell(date_row, col, week.start_date.toordinal() + offset - 693594)
            sheet.cell(date_row, col).number_format = "d-mmm"
            _label(sheet.cell(date_row, col))

        sheet.cell(week_row, 2, week.week_number)
        _label(sheet.cell(week_row, 2), light=True)
        for offset, day in enumerate(WEEKDAYS):
            col = 3 + offset
            sheet.cell(week_row, col, week.start_date.toordinal() + offset - 693594)
            sheet.cell(week_row, col).number_format = "ddd d-mmm"
            _label(sheet.cell(week_row, col), light=True)
        sheet.cell(week_row, 10, "Total")
        _label(sheet.cell(week_row, 10), light=True)

        row_labels = {
            type_row: "Type",
            plan_row: "Plan",
            planned_distance_row: "Distance (km)",
            actual_row: "Actual",
            actual_distance_row: "Distance (km)",
            remarks_row: "Remarks",
        }
        for row, label in row_labels.items():
            sheet.cell(row, 2, label)
            _label(sheet.cell(row, 2), light=True)

        for offset, session in enumerate(week.sessions):
            col = 3 + offset
            sheet.cell(type_row, col, session.session_type)
            sheet.cell(plan_row, col, session.plan)
            sheet.cell(planned_distance_row, col, session.planned_km)
            for row in [type_row, plan_row, planned_distance_row, actual_row, actual_distance_row, remarks_row]:
                _body(sheet.cell(row, col))
        sheet.cell(planned_distance_row, 10, f"=SUM(C{planned_distance_row}:I{planned_distance_row})")
        sheet.cell(actual_distance_row, 10, f"=SUM(C{actual_distance_row}:I{actual_distance_row})")
        _label(sheet.cell(planned_distance_row, 10), light=True)
        _label(sheet.cell(actual_distance_row, 10), light=True)

        side_items = [
            ("Phase", week.phase, "Target km", round(week.target_km, 1)),
            ("Focus", week.focus, "Long Run", week.long_run_summary),
            ("Key", week.key_sessions, "Date Range", f"{week.start_date:%d %b}-{week.end_date:%d %b}"),
            ("Notes", week.notes, "Adjust", week.adjust_note),
            ("Strength", week.strength_note, "Fuel", week.fuel_note),
            ("Risk", week.risk_note, "Race Fit", week.race_fit),
        ]
        for row_offset, (label1, value1, label2, value2) in enumerate(side_items, start=0):
            row = week_row + row_offset
            sheet.cell(row, 12, label1)
            sheet.cell(row, 13, value1)
            sheet.cell(row, 15, label2)
            sheet.cell(row, 16, value2)
            _label(sheet.cell(row, 12), light=True)
            _body(sheet.cell(row, 13))
            if label2:
                _label(sheet.cell(row, 15), light=True)
                _body(sheet.cell(row, 16))

        for row in range(start_row, start_row + 8):
            for col in range(2, 17):
                _border(sheet.cell(row, col))
        sheet.row_dimensions[plan_row].height = 48
        sheet.row_dimensions[start_row + 8].height = 12
        sheet.row_dimensions[start_row + 9].height = 12


def _label(cell, light: bool = False) -> None:
    cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN if light else GREEN)
    cell.font = Font(bold=True, color=TEXT if light else WHITE, size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    _border(cell)


def _body(cell) -> None:
    cell.font = Font(color=TEXT, size=10)
    cell.alignment = Alignment(wrap_text=True, vertical="center")
    _border(cell)


def _border(cell) -> None:
    side = Side(style="thin", color=BORDER)
    cell.border = Border(left=side, right=side, top=side, bottom=side)
